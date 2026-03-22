"""
=================================================================
MEDICAL DATA COLUMN ALIGNER
=================================================================
Aligns misaligned columns in medical professional databases
Supports multiple sheets, no-header detection, and medical patterns
=================================================================
"""

import pandas as pd
import numpy as np
import re
import os
import sys
import json
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from colorama import Fore, Back, Style, init
from tqdm import tqdm
from fuzzywuzzy import fuzz
import warnings

warnings.filterwarnings('ignore')
init(autoreset=True)

# =================================================================
# CONFIGURATION
# =================================================================

CONFIG = {
    'black_row_threshold': 0.7,  # 70% black cells to identify separator
    'sample_size': 20,  # Rows to analyze for pattern detection
    'min_match_confidence': 0.65,  # Minimum confidence for column matching
    'auto_backup': True,
    'preserve_formatting': True,
    'verbose': True,
    'max_empty_ratio': 0.8,  # Max 80% empty cells in column
    'fuzzy_match_threshold': 85,  # For text similarity matching
    'output_format': 'xlsx',  # xlsx or csv
    'black_color_codes': ['00000000', 'FF000000', '000000']
}

# =================================================================
# MEDICAL-SPECIFIC PATTERN LIBRARY
# =================================================================

MEDICAL_PATTERNS = {
    # Doctor Names with Credentials
    'doctor_name_full': {
        'priority': 20,
        'regex': [
            r'^(?:Dr\.?|Doctor|MD|DO)\s+[A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+(?:,?\s+(?:M\.?D\.?|D\.?O\.?|Ph\.?D\.?|MD|DO|PhD|MBBS|DDS|DMD))?',
            r'^[A-Z][a-z]+\s+[A-Z][a-z]+(?:,?\s+(?:M\.?D\.?|D\.?O\.?|Ph\.?D\.?|MD|DO|PhD|MBBS|DDS|DMD))',
        ],
        'keywords': ['dr', 'doctor', 'md', 'do', 'phd'],
        'examples': ['Dr. John Smith, MD', 'Sarah Johnson, DO', 'Dr. Michael Chen'],
        'validator': lambda x: bool(re.search(r'[A-Z][a-z]+.*[A-Z][a-z]+', str(x)))
    },
    
    'doctor_first_name': {
        'priority': 12,
        'regex': [r'^[A-Z][a-z]{2,20}$'],
        'keywords': ['first', 'name'],
        'examples': ['John', 'Sarah', 'Michael'],
        'validator': lambda x: isinstance(x, str) and 2 <= len(x) <= 20 and x.isalpha()
    },
    
    'doctor_last_name': {
        'priority': 12,
        'regex': [r'^[A-Z][a-z]{2,30}$'],
        'keywords': ['last', 'surname', 'family'],
        'examples': ['Smith', 'Johnson', 'Chen'],
        'validator': lambda x: isinstance(x, str) and 2 <= len(x) <= 30 and x.isalpha()
    },
    
    'credentials': {
        'priority': 15,
        'regex': [
            r'^(?:M\.?D\.?|D\.?O\.?|Ph\.?D\.?|MBBS|DDS|DMD|DPM|OD|PharmD|PA-C|NP|RN|CNS|APRN)',
            r'^(?:MD|DO|PhD|MBBS|DDS|DMD|DPM|OD)(?:,\s*(?:FACP|FACS|FAAD|FACOG))?',
        ],
        'keywords': ['credential', 'degree', 'certification'],
        'examples': ['MD', 'DO', 'PhD', 'MD, FACP', 'MBBS'],
        'validator': lambda x: bool(re.search(r'^[A-Z]{2,6}', str(x).strip()))
    },
    
    # Medical Specialties
    'medical_specialty': {
        'priority': 18,
        'regex': [
            r'(?i)(neurology|cardiology|orthopedics|pediatrics|family medicine|internal medicine|psychiatry|dermatology|oncology|radiology|anesthesiology|emergency medicine|surgery|general surgery|obstetrics|gynecology|ophthalmology|otolaryngology|ENT|urology|nephrology|gastroenterology|endocrinology|rheumatology|pulmonology|infectious disease|pathology|physical medicine|rehabilitation|sports medicine|pain management|palliative care|hospice|geriatrics|allergy|immunology)',
        ],
        'keywords': ['specialty', 'specialization', 'department', 'practice'],
        'examples': ['Neurology', 'Family Medicine', 'Cardiology', 'Pediatrics'],
        'validator': lambda x: isinstance(x, str) and len(x) > 3 and not x.isdigit()
    },
    
    'subspecialty': {
        'priority': 14,
        'regex': [
            r'(?i)(interventional|invasive|non-invasive|minimally invasive|pediatric|adult|geriatric|acute|chronic|preventive|diagnostic)',
        ],
        'keywords': ['subspecialty', 'focus', 'area'],
        'examples': ['Interventional Cardiology', 'Pediatric Neurology'],
        'validator': lambda x: isinstance(x, str) and len(x) > 5
    },
    
    # Contact Information
    'phone_number': {
        'priority': 19,
        'regex': [
            r'^\+?1?\s*\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})$',
            r'^\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})$',
            r'^[0-9]{10}$',
            r'^\+[0-9]{1,3}\s*\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}$',
        ],
        'keywords': ['phone', 'tel', 'telephone', 'mobile', 'cell', 'contact'],
        'examples': ['(555) 123-4567', '555-123-4567', '5551234567', '+1-555-123-4567'],
        'validator': lambda x: len(re.sub(r'\D', '', str(x))) >= 10
    },
    
    'fax_number': {
        'priority': 16,
        'regex': [
            r'^\+?1?\s*\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})$',
        ],
        'keywords': ['fax', 'facsimile'],
        'examples': ['(555) 123-4568', '555-123-4568'],
        'validator': lambda x: len(re.sub(r'\D', '', str(x))) >= 10
    },
    
    'email': {
        'priority': 17,
        'regex': [
            r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
        ],
        'keywords': ['email', 'mail', 'contact'],
        'examples': ['doctor@hospital.com', 'john.smith@medicalcenter.org'],
        'validator': lambda x: '@' in str(x) and '.' in str(x).split('@')[-1]
    },
    
    # URLs
    'profile_url': {
        'priority': 19,
        'regex': [
            r'^https?://(?:www\.)?[\w\-\.]+/(?:doctors?|physicians?|providers?|profiles?|team|staff|directory)/[\w\-/]+',
            r'^https?://[\w\-\.]+/[\w\-/]*(?:profile|doctor|physician|provider)',
        ],
        'keywords': ['profile', 'page', 'url', 'link', 'webpage'],
        'examples': ['https://hospital.com/doctors/john-smith', 'https://clinic.org/providers/sarah-jones'],
        'validator': lambda x: str(x).startswith('http') and any(kw in str(x).lower() for kw in ['doctor', 'physician', 'provider', 'profile', 'staff'])
    },
    
    'image_url': {
        'priority': 16,
        'regex': [
            r'^https?://.*\.(jpg|jpeg|png|gif|webp|svg)',
            r'^https?://.*(?:image|photo|picture|img)',
        ],
        'keywords': ['image', 'photo', 'picture', 'img', 'avatar'],
        'examples': ['https://hospital.com/images/doctor123.jpg', 'https://cdn.clinic.org/photos/doc.png'],
        'validator': lambda x: str(x).startswith('http') and (any(ext in str(x).lower() for ext in ['.jpg', '.jpeg', '.png', '.gif', '.webp']) or 'image' in str(x).lower())
    },
    
    'website_url': {
        'priority': 14,
        'regex': [
            r'^https?://(?:www\.)?[\w\-\.]+\.[a-z]{2,}',
        ],
        'keywords': ['website', 'url', 'link', 'web'],
        'examples': ['https://www.hospital.com', 'http://clinic.org'],
        'validator': lambda x: str(x).startswith('http')
    },
    
    # Address Components
    'full_address': {
        'priority': 17,
        'regex': [
            r'\d+\s+[\w\s]+(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|Drive|Dr|Court|Ct|Circle|Way|Parkway|Pkwy)',
            r'\d+\s+[A-Z][\w\s]+,\s*[A-Z][\w\s]*,\s*[A-Z]{2}\s*\d{5}',
        ],
        'keywords': ['address', 'location', 'street'],
        'examples': ['123 Main Street, Suite 200, Boston, MA 02115', '456 Oak Avenue, New York, NY 10001'],
        'validator': lambda x: bool(re.search(r'\d+.*(?:street|st|avenue|ave|road|rd)', str(x), re.IGNORECASE))
    },
    
    'street_address': {
        'priority': 16,
        'regex': [
            r'^\d+\s+[\w\s]+(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|Drive|Dr|Court|Ct|Circle|Way|Parkway|Pkwy)\.?',
        ],
        'keywords': ['street', 'address1', 'address'],
        'examples': ['123 Main Street', '456 Oak Avenue', '789 Elm Boulevard'],
        'validator': lambda x: bool(re.search(r'^\d+', str(x))) and len(str(x)) > 5
    },
    
    'suite_number': {
        'priority': 13,
        'regex': [
            r'^(?:Suite|Ste|Unit|#)\s*[A-Z0-9\-]+',
            r'^[A-Z0-9]{1,5}$',
        ],
        'keywords': ['suite', 'unit', 'apartment', 'apt', 'office'],
        'examples': ['Suite 200', 'Ste 4B', 'Unit 305', '#12A'],
        'validator': lambda x: isinstance(x, str) and len(x) <= 10
    },
    
    'city': {
        'priority': 15,
        'regex': [
            r'^[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*$',
        ],
        'keywords': ['city', 'town', 'municipality'],
        'examples': ['Boston', 'New York', 'Los Angeles', 'San Francisco'],
        'validator': lambda x: isinstance(x, str) and 2 <= len(x) <= 30 and not any(c.isdigit() for c in str(x))
    },
    
    'state': {
        'priority': 16,
        'regex': [
            r'^[A-Z]{2}$',
            r'^(Alabama|Alaska|Arizona|Arkansas|California|Colorado|Connecticut|Delaware|Florida|Georgia|Hawaii|Idaho|Illinois|Indiana|Iowa|Kansas|Kentucky|Louisiana|Maine|Maryland|Massachusetts|Michigan|Minnesota|Mississippi|Missouri|Montana|Nebraska|Nevada|New Hampshire|New Jersey|New Mexico|New York|North Carolina|North Dakota|Ohio|Oklahoma|Oregon|Pennsylvania|Rhode Island|South Carolina|South Dakota|Tennessee|Texas|Utah|Vermont|Virginia|Washington|West Virginia|Wisconsin|Wyoming)$',
        ],
        'keywords': ['state', 'province'],
        'examples': ['MA', 'NY', 'CA', 'TX', 'Massachusetts'],
        'validator': lambda x: len(str(x)) == 2 and str(x).isupper() or str(x) in ['Massachusetts', 'California', 'New York']
    },
    
    'zip_code': {
        'priority': 18,
        'regex': [
            r'^\d{5}(?:-\d{4})?$',
        ],
        'keywords': ['zip', 'zipcode', 'postal'],
        'examples': ['02115', '10001', '90210', '12345-6789'],
        'validator': lambda x: len(re.sub(r'\D', '', str(x))) in [5, 9]
    },
    
    # Ratings and Reviews
    'rating': {
        'priority': 15,
        'regex': [
            r'^[0-5](?:\.[0-9]{1,2})?$',
            r'^[0-5](?:\.[0-9]{1,2})?(?:/5)?(?:\s*(?:stars?|★))?',
        ],
        'keywords': ['rating', 'score', 'stars'],
        'examples': ['4.5', '5.0', '3.8/5', '4.2 stars', '★★★★☆'],
        'validator': lambda x: bool(re.search(r'^[0-5]\.?[0-9]?', str(x).strip('★☆ ')))
    },
    
    'review_count': {
        'priority': 13,
        'regex': [
            r'^\d+\s*(?:reviews?|ratings?)?$',
        ],
        'keywords': ['reviews', 'review count', 'number of reviews'],
        'examples': ['234 reviews', '15', '1,234 ratings'],
        'validator': lambda x: str(x).replace(',', '').replace(' reviews', '').replace(' ratings', '').isdigit()
    },
    
    # Medical Organization Info
    'hospital_name': {
        'priority': 16,
        'regex': [
            r'(?i)[\w\s]+(?:Hospital|Medical Center|Clinic|Health System|Healthcare|Medical Group|Physicians Group|Associates)',
        ],
        'keywords': ['hospital', 'clinic', 'medical center', 'practice', 'group'],
        'examples': ['Massachusetts General Hospital', 'Mayo Clinic', 'Cleveland Clinic'],
        'validator': lambda x: any(kw in str(x).lower() for kw in ['hospital', 'clinic', 'medical', 'health', 'physicians'])
    },
    
    'department': {
        'priority': 14,
        'regex': [
            r'(?i)(?:Department of|Dept\.|Division of)\s+[\w\s]+',
        ],
        'keywords': ['department', 'division', 'section'],
        'examples': ['Department of Neurology', 'Cardiology Division'],
        'validator': lambda x: 'department' in str(x).lower() or 'division' in str(x).lower()
    },
    
    # Appointment and Action Phrases
    'appointment_text': {
        'priority': 12,
        'regex': [
            r'(?i)(?:schedule|book|make|request)\s+(?:an?\s+)?(?:appointment|consultation|visit)',
            r'(?i)(?:call|contact)\s+(?:to|for)\s+(?:schedule|book)',
            r'(?i)accepting\s+new\s+patients',
        ],
        'keywords': ['appointment', 'schedule', 'book', 'consultation'],
        'examples': ['Schedule Appointment', 'Book Now', 'Request Consultation', 'Accepting New Patients'],
        'validator': lambda x: any(kw in str(x).lower() for kw in ['appointment', 'schedule', 'book', 'consult', 'accepting'])
    },
    
    'action_button': {
        'priority': 10,
        'regex': [
            r'(?i)^(?:view profile|see details|more info|learn more|read more|click here|visit page)$',
        ],
        'keywords': ['view', 'see', 'more', 'click', 'visit'],
        'examples': ['View Profile', 'See Details', 'More Info', 'Learn More'],
        'validator': lambda x: isinstance(x, str) and len(x) < 30 and any(kw in str(x).lower() for kw in ['view', 'see', 'more', 'click'])
    },
    
    # Additional Medical Info
    'npi_number': {
        'priority': 17,
        'regex': [
            r'^(?:NPI:?\s*)?[0-9]{10}$',
        ],
        'keywords': ['npi', 'national provider identifier'],
        'examples': ['1234567890', 'NPI: 9876543210'],
        'validator': lambda x: len(re.sub(r'\D', '', str(x))) == 10
    },
    
    'license_number': {
        'priority': 15,
        'regex': [
            r'^[A-Z]{1,3}[0-9]{5,10}$',
        ],
        'keywords': ['license', 'licence', 'medical license'],
        'examples': ['MD123456', 'CA987654321'],
        'validator': lambda x: bool(re.search(r'^[A-Z]+[0-9]+$', str(x)))
    },
    
    'years_experience': {
        'priority': 13,
        'regex': [
            r'^\d+\s*(?:years?|yrs?)(?:\s+(?:of\s+)?(?:experience|practice))?',
        ],
        'keywords': ['experience', 'years', 'practice'],
        'examples': ['15 years', '20 years experience', '5 yrs'],
        'validator': lambda x: bool(re.search(r'\d+', str(x))) and 'year' in str(x).lower()
    },
    
    'board_certification': {
        'priority': 14,
        'regex': [
            r'(?i)board\s+certified(?:\s+in\s+[\w\s]+)?',
            r'(?i)certified\s+by\s+[\w\s]+',
        ],
        'keywords': ['board certified', 'certification', 'certified'],
        'examples': ['Board Certified', 'Board Certified in Neurology', 'Certified by ABMS'],
        'validator': lambda x: 'certif' in str(x).lower()
    },
    
    'education': {
        'priority': 13,
        'regex': [
            r'(?i)[\w\s]+(?:University|College|School of Medicine|Medical School)',
        ],
        'keywords': ['education', 'medical school', 'university', 'college'],
        'examples': ['Harvard Medical School', 'Johns Hopkins University'],
        'validator': lambda x: any(kw in str(x).lower() for kw in ['university', 'college', 'school', 'medical'])
    },
    
    'languages': {
        'priority': 12,
        'regex': [
            r'(?i)(?:English|Spanish|French|German|Chinese|Mandarin|Hindi|Arabic|Portuguese|Russian|Japanese|Korean)(?:,\s*(?:English|Spanish|French|German|Chinese|Mandarin|Hindi|Arabic|Portuguese|Russian|Japanese|Korean))*',
        ],
        'keywords': ['language', 'languages spoken', 'speaks'],
        'examples': ['English, Spanish', 'English, Mandarin, Spanish'],
        'validator': lambda x: any(lang in str(x) for lang in ['English', 'Spanish', 'French', 'Chinese', 'Hindi'])
    },
    
    'insurance_accepted': {
        'priority': 11,
        'regex': [
            r'(?i)(?:accepts?|accepting)\s+(?:most\s+)?(?:insurance|plans)',
            r'(?i)(?:Aetna|Blue Cross|Cigna|UnitedHealthcare|Medicare|Medicaid|Humana)',
        ],
        'keywords': ['insurance', 'plans accepted', 'coverage'],
        'examples': ['Accepts most insurance', 'Aetna, Blue Cross, Cigna'],
        'validator': lambda x: 'insurance' in str(x).lower() or any(ins in str(x) for ins in ['Aetna', 'Cigna', 'Medicare'])
    },
    
    'gender': {
        'priority': 12,
        'regex': [
            r'^(?:Male|Female|M|F|Non-binary)$',
        ],
        'keywords': ['gender', 'sex'],
        'examples': ['Male', 'Female', 'M', 'F'],
        'validator': lambda x: str(x).strip() in ['Male', 'Female', 'M', 'F', 'Non-binary']
    },
    
    # Dates and Times
    'date': {
        'priority': 11,
        'regex': [
            r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}',
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',
        ],
        'keywords': ['date', 'since', 'joined'],
        'examples': ['01/15/2020', '2020-01-15'],
        'validator': lambda x: bool(re.search(r'\d{1,4}[-/]\d{1,2}', str(x)))
    },
    
    'office_hours': {
        'priority': 10,
        'regex': [
            r'(?i)(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun).*\d{1,2}:\d{2}',
            r'\d{1,2}:\d{2}\s*(?:AM|PM|am|pm)',
        ],
        'keywords': ['hours', 'office hours', 'available'],
        'examples': ['Mon-Fri: 9:00 AM - 5:00 PM', '9:00 AM - 5:00 PM'],
        'validator': lambda x: bool(re.search(r'\d{1,2}:\d{2}', str(x)))
    },
    
    # Generic Patterns
    'number': {
        'priority': 5,
        'regex': [
            r'^-?\d+(?:\.\d+)?$',
        ],
        'keywords': ['number', 'count', 'id'],
        'examples': ['123', '45.67', '-89'],
        'validator': lambda x: str(x).replace('.', '').replace('-', '').isdigit()
    },
    
    'text_short': {
        'priority': 3,
        'regex': [
            r'^.{1,50}$',
        ],
        'keywords': [],
        'examples': ['Short text content'],
        'validator': lambda x: isinstance(x, str) and 1 <= len(x) <= 50
    },
    
    'text_long': {
        'priority': 2,
        'regex': [
            r'^.{51,}$',
        ],
        'keywords': ['description', 'bio', 'about', 'notes'],
        'examples': ['Long form text content with multiple sentences...'],
        'validator': lambda x: isinstance(x, str) and len(x) > 50
    },
}

# =================================================================
# LOGGING SETUP
# =================================================================

def setup_logging():
    """Configure logging with file and console handlers"""
    log_dir = Path('logs')
    log_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = log_dir / f'alignment_{timestamp}.log'
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(levelname)8s | %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(__name__)

logger = setup_logging()

# =================================================================
# HELPER FUNCTIONS
# =================================================================

def print_header(text):
    """Print formatted header"""
    print(f"\n{Fore.CYAN}{'='*70}")
    print(f"{Fore.CYAN}{text.center(70)}")
    print(f"{Fore.CYAN}{'='*70}{Style.RESET_ALL}\n")

def print_success(text):
    """Print success message"""
    print(f"{Fore.GREEN}✓ {text}{Style.RESET_ALL}")

def print_error(text):
    """Print error message"""
    print(f"{Fore.RED}✗ {text}{Style.RESET_ALL}")

def print_warning(text):
    """Print warning message"""
    print(f"{Fore.YELLOW}⚠ {text}{Style.RESET_ALL}")

def print_info(text):
    """Print info message"""
    print(f"{Fore.BLUE}ℹ {text}{Style.RESET_ALL}")

# =================================================================
# CORE CLASSES
# =================================================================

class MedicalColumnAligner:
    """Main class for aligning medical data columns"""
    
    def __init__(self, input_file, config=None):
        self.input_file = Path(input_file)
        self.config = config or CONFIG
        self.patterns = MEDICAL_PATTERNS
        self.workbook = None
        self.results = []
        
        # Validate input
        if not self.input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        if self.input_file.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Unsupported file format: {self.input_file.suffix}")
        
        logger.info(f"Initialized aligner for: {self.input_file}")
    
    def detect_pattern(self, value, column_values=None):
        """Detect the pattern type of a value"""
        if pd.isna(value) or str(value).strip() == '':
            return None
        
        value_str = str(value).strip()
        matches = {}
        
        for pattern_name, pattern_config in self.patterns.items():
            score = 0
            
            # Test regex patterns
            for regex in pattern_config['regex']:
                try:
                    if re.search(regex, value_str, re.IGNORECASE):
                        score += pattern_config['priority']
                        break
                except:
                    continue
            
            # Test validator function
            try:
                if pattern_config['validator'](value):
                    score += pattern_config['priority'] * 0.5
            except:
                pass
            
            if score > 0:
                matches[pattern_name] = score
        
        if not matches:
            return 'text_short' if len(value_str) <= 50 else 'text_long'
        
        # Return highest scoring pattern
        return max(matches, key=matches.get)
    
    def detect_column_type(self, column_values, column_name=''):
        """Detect the type of an entire column"""
        # Filter out empty values
        non_empty = [v for v in column_values if pd.notna(v) and str(v).strip() != '']
        
        if not non_empty:
            return 'text_short'
        
        # Sample values for analysis
        sample = non_empty[:self.config['sample_size']]
        pattern_counts = {}
        
        for value in sample:
            pattern = self.detect_pattern(value)
            if pattern:
                pattern_counts[pattern] = pattern_counts.get(pattern, 0) + 1
        
        if not pattern_counts:
            return 'text_short'
        
        # Calculate confidence scores
        total_samples = len(sample)
        scores = {}
        
        for pattern, count in pattern_counts.items():
            confidence = count / total_samples
            priority = self.patterns[pattern]['priority']
            scores[pattern] = confidence * priority
            
            # Bonus for matching column name
            if column_name:
                for keyword in self.patterns[pattern]['keywords']:
                    if keyword in column_name.lower():
                        scores[pattern] += 5
        
        best_pattern = max(scores, key=scores.get)
        confidence = pattern_counts[best_pattern] / total_samples
        
        logger.debug(f"Column '{column_name}' detected as '{best_pattern}' (confidence: {confidence:.2f})")
        
        return best_pattern
    
    def find_black_rows(self, sheet_name):
        """Find rows marked with black background color"""
        try:
            wb = load_workbook(self.input_file, data_only=True)
            ws = wb[sheet_name]
            black_rows = []
            
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                black_cells = 0
                total_cells = 0
                
                for cell in row:
                    total_cells += 1
                    if cell.fill and cell.fill.start_color:
                        color = cell.fill.start_color.rgb
                        if color in self.config['black_color_codes']:
                            black_cells += 1
                
                if total_cells > 0 and (black_cells / total_cells) >= self.config['black_row_threshold']:
                    black_rows.append(row_idx)
            
            wb.close()
            return black_rows
            
        except Exception as e:
            logger.error(f"Error finding black rows in '{sheet_name}': {e}")
            return []
    
    def split_into_segments(self, df, black_rows):
        """Split dataframe into segments based on black rows"""
        segments = []
        start_idx = 0
        
        for black_row in black_rows:
            # Convert to 0-based index
            end_idx = black_row - 1
            
            if start_idx < end_idx:
                segment = df.iloc[start_idx:end_idx].copy()
                if not segment.empty:
                    segments.append({
                        'start_row': start_idx + 1,
                        'end_row': end_idx,
                        'data': segment,
                        'is_separator': False
                    })
            
            # Add black row as separator
            if black_row - 1 < len(df):
                segments.append({
                    'start_row': black_row,
                    'end_row': black_row,
                    'data': df.iloc[black_row-1:black_row].copy(),
                    'is_separator': True
                })
            
            start_idx = black_row
        
        # Add remaining rows
        if start_idx < len(df):
            segment = df.iloc[start_idx:].copy()
            if not segment.empty:
                segments.append({
                    'start_row': start_idx + 1,
                    'end_row': len(df),
                    'data': segment,
                    'is_separator': False
                })
        
        return segments
    
    def analyze_reference_structure(self, segment_data):
        """Analyze a reference segment to determine column structure"""
        structure = []
        num_cols = len(segment_data.columns)
        
        for col_idx in range(num_cols):
            column_values = segment_data.iloc[:, col_idx].tolist()
            column_name = f"Column_{col_idx}"
            
            # Detect column type
            detected_type = self.detect_column_type(column_values, column_name)
            
            # Get sample values
            samples = [v for v in column_values if pd.notna(v) and str(v).strip() != ''][:3]
            
            structure.append({
                'index': col_idx,
                'name': column_name,
                'type': detected_type,
                'priority': self.patterns[detected_type]['priority'],
                'samples': samples
            })
        
        # Sort by priority (descending)
        structure.sort(key=lambda x: x['priority'], reverse=True)
        
        return structure
    
    def align_segment_to_reference(self, segment_data, reference_structure):
        """Align a segment to match reference structure"""
        num_ref_cols = len(reference_structure)
        aligned_rows = []
        
        for row_idx in range(len(segment_data)):
            row = segment_data.iloc[row_idx].tolist()
            new_row = [None] * num_ref_cols
            used_indices = []
            
            # First pass: Match by type with high confidence
            for ref_idx, ref_col in enumerate(reference_structure):
                best_match_idx = None
                best_score = 0
                
                for data_idx, value in enumerate(row):
                    if data_idx in used_indices:
                        continue
                    
                    if pd.isna(value) or str(value).strip() == '':
                        continue
                    
                    score = 0
                    detected_type = self.detect_pattern(value)
                    
                    # Type match
                    if detected_type == ref_col['type']:
                        score += ref_col['priority'] * 2
                    
                    # Position proximity bonus
                    position_diff = abs(data_idx - ref_col['index'])
                    if position_diff == 0:
                        score += 10
                    elif position_diff == 1:
                        score += 5
                    elif position_diff == 2:
                        score += 2
                    
                    # Fuzzy match with samples
                    if isinstance(value, str) and ref_col['samples']:
                        for sample in ref_col['samples']:
                            if isinstance(sample, str):
                                similarity = fuzz.ratio(str(value).lower(), str(sample).lower())
                                if similarity >= self.config['fuzzy_match_threshold']:
                                    score += similarity / 10
                    
                    if score > best_score:
                        best_score = score
                        best_match_idx = data_idx
                
                # Assign if confidence threshold met
                if best_match_idx is not None and best_score >= self.config['min_match_confidence'] * 10:
                    new_row[ref_idx] = row[best_match_idx]
                    used_indices.append(best_match_idx)
            
            # Second pass: Fill remaining by position
            for ref_idx, ref_col in enumerate(reference_structure):
                if new_row[ref_idx] is None:
                    original_idx = ref_col['index']
                    if original_idx < len(row) and original_idx not in used_indices:
                        if pd.notna(row[original_idx]) and str(row[original_idx]).strip() != '':
                            new_row[ref_idx] = row[original_idx]
                            used_indices.append(original_idx)
            
            aligned_rows.append(new_row)
        
        # Create DataFrame with proper column names
        column_names = [f"{ref['type']}_{ref['index']}" for ref in reference_structure]
        aligned_df = pd.DataFrame(aligned_rows, columns=column_names)
        
        return aligned_df
    
    def process_sheet(self, sheet_name):
        """Process a single sheet"""
        try:
            print_info(f"Processing sheet: {sheet_name}")
            logger.info(f"Processing sheet: {sheet_name}")
            
            # Read sheet data
            df = pd.read_excel(self.input_file, sheet_name=sheet_name, header=None)
            
            if df.empty:
                print_warning(f"Sheet '{sheet_name}' is empty")
                return None
            
            # Find black rows
            black_rows = self.find_black_rows(sheet_name)
            
            if not black_rows:
                print_warning(f"No black separator rows found in '{sheet_name}'")
                # Process without segments
                reference_structure = self.analyze_reference_structure(df)
                aligned_df = self.align_segment_to_reference(df, reference_structure)
                return {
                    'sheet_name': sheet_name,
                    'data': aligned_df,
                    'structure': reference_structure,
                    'segments': 1,
                    'black_rows': 0
                }
            
            print_success(f"Found {len(black_rows)} black separator row(s)")
            
            # Split into segments
            segments = self.split_into_segments(df, black_rows)
            print_info(f"Split into {len(segments)} segment(s)")
            
            # Find first valid reference segment
            reference_segment = None
            for seg in segments:
                if not seg['is_separator'] and not seg['data'].empty:
                    reference_segment = seg
                    break
            
            if not reference_segment:
                print_error(f"No valid reference segment found in '{sheet_name}'")
                return None
            
            # Analyze reference structure
            reference_structure = self.analyze_reference_structure(reference_segment['data'])
            
            print_info(f"Detected {len(reference_structure)} column types:")
            for col in reference_structure[:5]:  # Show first 5
                print(f"  - {col['name']}: {col['type']} (priority: {col['priority']})")
            if len(reference_structure) > 5:
                print(f"  ... and {len(reference_structure) - 5} more")
            
            # Align all segments
            aligned_segments = []
            
            for seg_idx, seg in enumerate(tqdm(segments, desc="Aligning segments", unit="segment")):
                if seg['is_separator']:
                    # Keep separator as-is
                    aligned_segments.append(seg['data'])
                else:
                    aligned_data = self.align_segment_to_reference(seg['data'], reference_structure)
                    aligned_segments.append(aligned_data)
            
            # Combine all segments
            final_df = pd.concat(aligned_segments, ignore_index=True)
            
            print_success(f"Aligned {len(segments)} segment(s) successfully")
            
            return {
                'sheet_name': sheet_name,
                'data': final_df,
                'structure': reference_structure,
                'segments': len(segments),
                'black_rows': len(black_rows)
            }
            
        except Exception as e:
            print_error(f"Error processing sheet '{sheet_name}': {e}")
            logger.error(f"Error processing sheet '{sheet_name}': {e}", exc_info=True)
            return None
    
    def process_all_sheets(self):
        """Process all sheets in the workbook"""
        print_header("MEDICAL DATA COLUMN ALIGNER")
        print_info(f"Input file: {self.input_file}")
        
        # Create backup
        if self.config['auto_backup']:
            backup_path = self.create_backup()
            print_success(f"Backup created: {backup_path}")
        
        # Get all sheet names
        try:
            xl_file = pd.ExcelFile(self.input_file)
            sheet_names = xl_file.sheet_names
            print_info(f"Found {len(sheet_names)} sheet(s)")
        except Exception as e:
            print_error(f"Error reading file: {e}")
            return None
        
        # Process each sheet
        processed_sheets = []
        
        for sheet_name in sheet_names:
            print(f"\n{Fore.YELLOW}{'─'*70}{Style.RESET_ALL}")
            result = self.process_sheet(sheet_name)
            
            if result:
                processed_sheets.append(result)
                self.results.append({
                    'sheet': sheet_name,
                    'status': 'success',
                    'segments': result['segments'],
                    'black_rows': result['black_rows']
                })
            else:
                self.results.append({
                    'sheet': sheet_name,
                    'status': 'failed',
                    'segments': 0,
                    'black_rows': 0
                })
        
        # Save results
        if processed_sheets:
            output_path = self.save_results(processed_sheets)
            print_success(f"\nOutput saved to: {output_path}")
            
            # Save structure report
            report_path = self.save_structure_report(processed_sheets)
            print_success(f"Structure report saved to: {report_path}")
            
            return output_path
        else:
            print_error("No sheets were successfully processed")
            return None
    
    def create_backup(self):
        """Create a backup of the original file"""
        backup_dir = Path('backups')
        backup_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{self.input_file.stem}_backup_{timestamp}{self.input_file.suffix}"
        backup_path = backup_dir / backup_name
        
        import shutil
        shutil.copy2(self.input_file, backup_path)
        
        logger.info(f"Backup created: {backup_path}")
        return backup_path
    
    def save_results(self, processed_sheets):
        """Save processed data to output file"""
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"{self.input_file.stem}_aligned_{timestamp}.xlsx"
        output_path = output_dir / output_name
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_result in processed_sheets:
                sheet_result['data'].to_excel(
                    writer,
                    sheet_name=sheet_result['sheet_name'][:31],  # Excel limit
                    index=False
                )
        
        # Apply formatting
        self.apply_formatting(output_path, processed_sheets)
        
        logger.info(f"Results saved to: {output_path}")
        return output_path
    
    def apply_formatting(self, output_path, processed_sheets):
        """Apply formatting to output file"""
        try:
            wb = load_workbook(output_path)
            
            for sheet_result in processed_sheets:
                sheet_name = sheet_result['sheet_name'][:31]
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                # Format header row
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            wb.close()
            
        except Exception as e:
            logger.warning(f"Could not apply formatting: {e}")
    
    def save_structure_report(self, processed_sheets):
        """Save detailed structure analysis report"""
        output_dir = Path('output')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_path = output_dir / f"structure_report_{timestamp}.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("MEDICAL DATA STRUCTURE ANALYSIS REPORT\n")
            f.write("="*70 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Input File: {self.input_file}\n\n")
            
            for sheet_result in processed_sheets:
                f.write(f"\n{'─'*70}\n")
                f.write(f"Sheet: {sheet_result['sheet_name']}\n")
                f.write(f"{'─'*70}\n")
                f.write(f"Segments: {sheet_result['segments']}\n")
                f.write(f"Black Rows: {sheet_result['black_rows']}\n\n")
                
                f.write("Detected Column Structure:\n")
                f.write(f"{'─'*70}\n")
                
                for col in sheet_result['structure']:
                    f.write(f"\nColumn {col['index'] + 1}:\n")
                    f.write(f"  Type: {col['type']}\n")
                    f.write(f"  Priority: {col['priority']}\n")
                    f.write(f"  Sample Values:\n")
                    for sample in col['samples']:
                        f.write(f"    - {sample}\n")
        
        logger.info(f"Structure report saved: {report_path}")
        return report_path
    
    def print_summary(self):
        """Print processing summary"""
        print_header("PROCESSING SUMMARY")
        
        success_count = sum(1 for r in self.results if r['status'] == 'success')
        failed_count = len(self.results) - success_count
        
        print(f"{Fore.CYAN}Total Sheets:{Style.RESET_ALL} {len(self.results)}")
        print(f"{Fore.GREEN}Successful:{Style.RESET_ALL} {success_count}")
        print(f"{Fore.RED}Failed:{Style.RESET_ALL} {failed_count}")
        
        if success_count > 0:
            print(f"\n{Fore.YELLOW}Details:{Style.RESET_ALL}")
            for result in self.results:
                if result['status'] == 'success':
                    status_icon = f"{Fore.GREEN}✓{Style.RESET_ALL}"
                    print(f"  {status_icon} {result['sheet']}: {result['segments']} segments, {result['black_rows']} separators")
                else:
                    status_icon = f"{Fore.RED}✗{Style.RESET_ALL}"
                    print(f"  {status_icon} {result['sheet']}: Failed")

# =================================================================
# MAIN EXECUTION
# =================================================================

def main():
    """Main execution function"""
    print_header("MEDICAL DATA COLUMN ALIGNER v1.0")
    
    # Check for input file
    input_dir = Path('input')
    input_dir.mkdir(exist_ok=True)
    
    excel_files = list(input_dir.glob('*.xlsx')) + list(input_dir.glob('*.xls'))
    
    if not excel_files:
        print_error("No Excel files found in 'input' folder")
        print_info("Please place your Excel file in the 'input' folder and run again")
        return
    
    if len(excel_files) > 1:
        print_info("Multiple Excel files found. Select file to process:")
        for idx, file in enumerate(excel_files, 1):
            print(f"  {idx}. {file.name}")
        
        while True:
            try:
                choice = int(input(f"\n{Fore.YELLOW}Enter file number: {Style.RESET_ALL}"))
                if 1 <= choice <= len(excel_files):
                    input_file = excel_files[choice - 1]
                    break
                else:
                    print_error("Invalid selection")
            except ValueError:
                print_error("Please enter a number")
    else:
        input_file = excel_files[0]
    
    print_info(f"Selected file: {input_file.name}")
    
    # Initialize aligner
    try:
        aligner = MedicalColumnAligner(input_file, CONFIG)
        
        # Process all sheets
        output_path = aligner.process_all_sheets()
        
        # Print summary
        aligner.print_summary()
        
        if output_path:
            print(f"\n{Fore.GREEN}{'='*70}")
            print(f"{Fore.GREEN}✓ PROCESSING COMPLETE!")
            print(f"{Fore.GREEN}{'='*70}{Style.RESET_ALL}")
            print(f"\n{Fore.CYAN}Output file:{Style.RESET_ALL} {output_path}")
            print(f"{Fore.CYAN}Log file:{Style.RESET_ALL} logs/alignment_*.log")
        
    except Exception as e:
        print_error(f"Fatal error: {e}")
        logger.error(f"Fatal error: {e}", exc_info=True)
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())